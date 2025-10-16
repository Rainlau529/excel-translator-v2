# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, jsonify, send_file, Response
import requests
import time
import os
import tempfile
import shutil
import threading
import uuid
import json
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 64 * 1024 * 1024  # 64 MB

# 任务状态存储
TASKS = {}
TASKS_LOCK = threading.Lock()
TASK_RETENTION_SECONDS = 1800

def translate_text(text: str) -> str:
    """谷歌翻译（保持不变）"""
    try:
        url = "https://translate.googleapis.com/translate_a/single"
        params = {"client": "gtx", "sl": "auto", "tl": "zh", "dt": "t", "q": text}
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        return "".join(seg[0] for seg in data[0]) if data and data[0] else "翻译失败"
    except Exception as e:
        print(f"[ERROR] translate_text: {e}")
        return "翻译失败"

# ---------- 任务管理 ----------
def _safe_update(task_id: str, kv: dict):
    with TASKS_LOCK:
        st = TASKS.get(task_id)
        if st:
            st.update(kv)

def _get_state(task_id: str):
    with TASKS_LOCK:
        return dict(TASKS.get(task_id) or {})

def _schedule_state_cleanup(task_id: str):
    def _cleanup():
        time.sleep(TASK_RETENTION_SECONDS)
        with TASKS_LOCK:
            TASKS.pop(task_id, None)
    threading.Thread(target=_cleanup, daemon=True).start()

def _run_task(file_path: str, task_id: str):
    _safe_update(task_id, {"status": "running", "message": "读取文件...", "percent": 0, "started_at": time.time()})
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        if not ws:
            raise ValueError("空工作表")

        # 找 Title 列
        title_col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if str(cell.value).strip().lower() == "title":
                title_col_idx = idx
                break
        if not title_col_idx:
            raise ValueError("找不到 Title 列")

        # 插入中文列
        insert_idx = title_col_idx + 1
        ws.insert_cols(insert_idx)
        ws.cell(row=1, column=insert_idx, value="中文")

        rows_to_translate = [
            r for r in range(2, ws.max_row + 1)
            if str(ws.cell(row=r, column=title_col_idx).value).strip()
        ]
        total = len(rows_to_translate)
        _safe_update(task_id, {"total": total, "current": 0})

        if total == 0:
            _safe_update(task_id, {"status": "done", "percent": 100, "message": "无需翻译"})
            return

        start = time.time()
        for idx, row in enumerate(rows_to_translate, 1):
            if _get_state(task_id).get("cancel_requested"):
                _safe_update(task_id, {"status": "canceled", "message": "已取消"})
                return

            title_cell = ws.cell(row=row, column=title_col_idx)
            translated = translate_text(str(title_cell.value))
            ws.cell(row=row, column=insert_idx, value=translated)

            elapsed = max(time.time() - start, 1e-6)
            eta = int((total - idx) / (idx / elapsed)) if idx else 0
            _safe_update(task_id, {
                "current": idx,
                "percent": int(idx * 100 / total),
                "eta_seconds": eta,
                "message": f"翻译中({idx}/{total})"
            })
            time.sleep(0.3)

        output_filename = os.path.splitext(os.path.basename(file_path))[0] + "_中文翻译.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), output_filename)
        wb.save(output_path)

        _safe_update(task_id, {
            "status": "done",
            "percent": 100,
            "eta_seconds": 0,
            "message": f"翻译完成，共处理 {total} 行",
            "download_filename": output_filename,
            "finished_at": time.time()
        })
    except Exception as e:
        _safe_update(task_id, {"status": "error", "message": str(e), "finished_at": time.time()})
    finally:
        try:
            base = os.path.dirname(file_path)
            if os.path.isdir(base):
                shutil.rmtree(base, ignore_errors=True)
        except Exception:
            pass
        _schedule_state_cleanup(task_id)

def start_background_task(file_path: str, filename: str) -> str:
    task_id = uuid.uuid4().hex[:12]
    with TASKS_LOCK:
        TASKS[task_id] = {
            "status": "idle",
            "percent": 0,
            "message": "",
            "filename": filename,
            "total": 0,
            "current": 0,
            "cancel_requested": False,
            "tmp_dir": os.path.dirname(file_path)
        }
    threading.Thread(target=_run_task, args=(file_path, task_id), daemon=True).start()
    return task_id

# ---------- SSE ----------
def sse_progress(task_id: str):
    def event_stream():
        while True:
            state = _get_state(task_id)
            if not state:
                yield f"data: {json.dumps({'task_id': task_id, 'status': 'error', 'message': '任务不存在'})}\n\n"
                break
            payload = json.dumps({
                "task_id": task_id,
                "status": state.get("status"),
                "percent": state.get("percent"),
                "eta_seconds": state.get("eta_seconds"),
                "message": state.get("message"),
                "download_url": f"/download/{state.get('download_filename')}" if state.get("download_filename") else None,
                "filename": state.get("filename"),
                "started_at": state.get("started_at"),
                "finished_at": state.get("finished_at"),
                "duration_seconds": state.get("duration_seconds")
            }, ensure_ascii=False)
            yield f"event: progress\ndata: {payload}\n\n"
            if state.get("status") in ("done", "error", "canceled"):
                break
            time.sleep(0.2)
    return Response(event_stream(), mimetype="text/event-stream")

# ---------- 路由 ----------
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '没有选择文件'}), 400
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': '请上传Excel文件(.xlsx)'}), 400

    try:
        filename = secure_filename(file.filename)
        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, filename)
        file.save(file_path)
        task_id = start_background_task(file_path, filename)
        return jsonify({'success': True, 'task_id': task_id}), 202
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/progress/<task_id>')
def progress(task_id):
    return sse_progress(task_id)

@app.route('/tasks/<task_id>', methods=['GET'])
def task_status(task_id):
    st = _get_state(task_id)
    return jsonify(st) if st else (jsonify({'error': '任务不存在'}), 404)

@app.route('/tasks/<task_id>/cancel', methods=['POST'])
def task_cancel(task_id):
    st = _get_state(task_id)
    if not st:
        return jsonify({'error': '任务不存在'}), 404
    if st.get("status") in ("done", "error", "canceled"):
        return jsonify({'success': True, 'message': '任务已结束'}), 200
    _safe_update(task_id, {"cancel_requested": True, "message": "取消中..."})
    return jsonify({'success': True, 'message': '已请求取消'}), 202

@app.route('/download/<filename>')
def download_file(filename):
    file_path = os.path.join(tempfile.gettempdir(), filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, download_name=filename)
    return "文件不存在", 404

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000, threaded=True)
